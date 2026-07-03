import xml.etree.ElementTree as ET

import openpyxl
from django.db.models import F, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import now
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from historial.utils import registrar_actividad
from ordenes.models import Orden
from usuarios.models import Material, Usuario, Vehiculo
from usuarios.views import admin_required


@admin_required
def reportes_admin(request):
    # Estadísticas de Órdenes
    ordenes = Orden.objects.all()
    total = ordenes.count()
    pendientes = ordenes.filter(estado="pendiente").count()
    en_ruta = ordenes.filter(estado="en_ruta").count()
    entregadas = ordenes.filter(estado="entregado").count()

    # Calcular porcentajes
    pct_pendientes = (pendientes * 100 / total) if total > 0 else 0
    pct_en_ruta = (en_ruta * 100 / total) if total > 0 else 0
    pct_entregadas = (entregadas * 100 / total) if total > 0 else 0

    # Materiales con stock crítico
    stock_critico = Material.objects.filter(
        stock_info__cantidad_actual__lt=F("stock_info__stock_minimo")
    ).select_related("stock_info")

    context = {
        # Resumen de Órdenes
        "total_ordenes": total,
        "pendientes": pendientes,
        "en_ruta": en_ruta,
        "entregadas": entregadas,
        "pct_pendientes": pct_pendientes,
        "pct_en_ruta": pct_en_ruta,
        "pct_entregadas": pct_entregadas,
        # Resumen de Usuarios
        "total_clientes": Usuario.objects.filter(rol="cliente").count(),
        "total_conductores": Usuario.objects.filter(rol="conductor").count(),
        # Resumen de Inventario y Vehículos
        "total_materiales": Material.objects.count(),
        "total_vehiculos": Vehiculo.objects.count(),
        # Financiero
        "total_ingresos": ordenes.aggregate(total=Sum("precio"))["total"] or 0,
        # Stock Crítico
        "stock_critico": stock_critico,
    }
    return render(request, "reportes/lista.html", context)


@admin_required
def exportar_reporte_pdf(request, tipo):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    titulo = f"Reporte de {tipo.replace('_', ' ').capitalize()}"
    elements.append(Paragraph(titulo, styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(f"Fecha de generación: {now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"])
    )
    elements.append(Spacer(1, 24))

    def format_money(val):
        try:
            v = float(val)
            rounded = int(round(v))
            s = str(rounded)
            parts = []
            while s:
                parts.append(s[-3:])
                s = s[:-3]
            return ".".join(reversed(parts))
        except Exception:
            return "0"

    data = []
    if tipo == "clientes":
        data.append(["ID", "Nombre", "Correo", "Teléfono", "Estado"])
        for u in Usuario.objects.filter(rol="cliente"):
            data.append(
                [u.id, f"{u.nombres} {u.apellidos}", u.email, u.telefono or "N/A", u.estado]
            )

    elif tipo == "materiales":
        data.append(["ID", "Nombre", "Tipo", "Precio", "Stock"])
        for m in Material.objects.all().select_related("stock_info"):
            p = m.precio or 0
            precio_formateado = format_money(p)
            tipo_material = m.tipo or "N/A"
            data.append([m.id, m.nombre, tipo_material, precio_formateado, m.stock])

    elif tipo == "ventas":
        data.append(["ID", "Cliente", "Fecha", "Total", "Estado"])
        for o in Orden.objects.all().select_related("cliente__usuario"):
            p = o.precio or 0
            precio_formateado = format_money(p)
            fecha_str = o.fecha.strftime("%Y-%m-%d") if o.fecha else "N/A"
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            data.append([o.id, cliente_nombre, fecha_str, precio_formateado, o.estado])

    elif tipo == "pedidos":
        data.append(["ID", "Cliente", "Materiales", "Total", "Estado"])
        for o in (
            Orden.objects.all()
            .select_related("cliente__usuario")
            .prefetch_related("detalles__material")
        ):
            materiales = ", ".join(
                [f"{d.cantidad} x {d.material.nombre}" for d in o.detalles.all()]
            )
            p = o.precio or 0
            precio_formateado = format_money(p)
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            data.append(
                [
                    o.id,
                    cliente_nombre,
                    materiales[:50] + "..." if len(materiales) > 50 else materiales,
                    precio_formateado,
                    o.estado,
                ]
            )

    # Estilo de la tabla
    if data:
        t = Table(data)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        elements.append(t)
    else:
        elements.append(Paragraph("No hay datos disponibles para este reporte.", styles["Normal"]))

    doc.build(elements)

    registrar_actividad(request, "otro", "reportes", None, f"Reporte de {tipo} exportado a PDF")

    return response


@admin_required
def exportar_reporte_excel(request, tipo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo.capitalize()}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_aligned = Alignment(horizontal="center")

    def format_money_raw(val):
        try:
            return float(val)
        except Exception:
            return 0.0

    if tipo == "clientes":
        headers = ["ID", "Nombre", "Correo", "Teléfono", "Estado"]
        ws.append(headers)
        for u in Usuario.objects.filter(rol="cliente"):
            ws.append([u.id, f"{u.nombres} {u.apellidos}", u.email, u.telefono or "N/A", u.estado])

    elif tipo == "materiales":
        headers = ["ID", "Nombre", "Tipo", "Precio", "Stock"]
        ws.append(headers)
        for m in Material.objects.all().select_related("stock_info"):
            ws.append([m.id, m.nombre, m.tipo or "N/A", format_money_raw(m.precio), m.stock])

    elif tipo == "ventas":
        headers = ["ID", "Cliente", "Fecha", "Total", "Estado"]
        ws.append(headers)
        for o in Orden.objects.all().select_related("cliente__usuario"):
            fecha_str = o.fecha.strftime("%Y-%m-%d") if o.fecha else "N/A"
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            ws.append([o.id, cliente_nombre, fecha_str, format_money_raw(o.precio), o.estado])

    elif tipo == "pedidos":
        headers = ["ID", "Cliente", "Materiales", "Total", "Estado"]
        ws.append(headers)
        for o in (
            Orden.objects.all()
            .select_related("cliente__usuario")
            .prefetch_related("detalles__material")
        ):
            materiales = ", ".join(
                [f"{d.cantidad} x {d.material.nombre}" for d in o.detalles.all()]
            )
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            ws.append([o.id, cliente_nombre, materiales, format_money_raw(o.precio), o.estado])

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(response)

    registrar_actividad(request, "otro", "reportes", None, f"Reporte de {tipo} exportado a Excel")
    return response


@admin_required
def exportar_reporte_xml(request, tipo):
    root = ET.Element("reporte")
    root.set("tipo", tipo)
    root.set("fecha_generacion", now().strftime("%Y-%m-%d %H:%M:%S"))

    if tipo == "clientes":
        for u in Usuario.objects.filter(rol="cliente"):
            item = ET.SubElement(root, "cliente")
            ET.SubElement(item, "id").text = str(u.id)
            ET.SubElement(item, "nombre").text = f"{u.nombres} {u.apellidos}"
            ET.SubElement(item, "email").text = u.email
            ET.SubElement(item, "telefono").text = u.telefono or ""
            ET.SubElement(item, "estado").text = u.estado

    elif tipo == "materiales":
        for m in Material.objects.all().select_related("stock_info"):
            item = ET.SubElement(root, "material")
            ET.SubElement(item, "id").text = str(m.id)
            ET.SubElement(item, "nombre").text = m.nombre
            ET.SubElement(item, "tipo").text = m.tipo or ""
            ET.SubElement(item, "precio").text = str(m.precio or 0)
            ET.SubElement(item, "stock").text = str(m.stock)

    elif tipo == "ventas":
        for o in Orden.objects.all().select_related("cliente__usuario"):
            item = ET.SubElement(root, "venta")
            ET.SubElement(item, "id").text = str(o.id)
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            ET.SubElement(item, "cliente").text = cliente_nombre
            fecha_str = o.fecha.strftime("%Y-%m-%d") if o.fecha else "N/A"
            ET.SubElement(item, "fecha").text = fecha_str
            ET.SubElement(item, "total").text = str(o.precio or 0)
            ET.SubElement(item, "estado").text = o.estado

    elif tipo == "pedidos":
        for o in (
            Orden.objects.all()
            .select_related("cliente__usuario")
            .prefetch_related("detalles__material")
        ):
            item = ET.SubElement(root, "pedido")
            ET.SubElement(item, "id").text = str(o.id)
            cliente_nombre = (
                f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
                if (o.cliente and o.cliente.usuario)
                else "N/A"
            )
            ET.SubElement(item, "cliente").text = cliente_nombre
            ET.SubElement(item, "total").text = str(o.precio or 0)
            ET.SubElement(item, "estado").text = o.estado
            dets = ET.SubElement(item, "detalles")
            for d in o.detalles.all():
                det = ET.SubElement(dets, "detalle")
                ET.SubElement(det, "material").text = d.material.nombre
                ET.SubElement(det, "cantidad").text = str(d.cantidad)

    # Convertir a string indentado
    from xml.dom import minidom

    xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")

    response = HttpResponse(xmlstr, content_type="application/xml")
    response["Content-Disposition"] = (
        f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.xml"'
    )

    registrar_actividad(request, "otro", "reportes", None, f"Reporte de {tipo} exportado a XML")
    return response
