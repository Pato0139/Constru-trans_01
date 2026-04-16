import os
from io import BytesIO
from decimal import Decimal

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse, Http404
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from .models import Orden, Entrega
from apps.usuarios.models import Usuario, Vehiculo, Material
from historial.utils import registrar_actividad

def buscar_pedidos_admin(cliente_query=None, fecha_query=None):
    """
    Lógica unificada para buscar pedidos por cliente o fecha.
    """
    pedidos = Orden.objects.all().order_by("-fecha")
    
    if cliente_query:
        pedidos = pedidos.filter(
            Q(cliente__nombres__icontains=cliente_query) |
            Q(cliente__apellidos__icontains=cliente_query)
        )
    
    if fecha_query:
        pedidos = pedidos.filter(fecha__date=fecha_query)
        
    return pedidos

@login_required
def lista_pedidos_admin(request):
    cliente_query = request.GET.get('cliente')
    fecha_query = request.GET.get('fecha')
    
    pedidos = buscar_pedidos_admin(cliente_query, fecha_query)
        
    return render(request, "ordenes/lista.html", {
        "pedidos": pedidos,
        "cliente_query": cliente_query,
        "fecha_query": fecha_query
    })

@login_required
def ver_pedido_admin(request, id):
    orden = get_object_or_404(Orden, id=id)
    return render(request, "ordenes/detalle.html", {
        "orden": orden
    })

@login_required
def crear_entrega(request, orden_id):
    orden = get_object_or_404(Orden, id=orden_id)
    conductores = Usuario.objects.filter(rol="conductor")
    vehiculos = Vehiculo.objects.all()

    if request.method == "POST":
        conductor_id = request.POST.get("conductor")
        vehiculo_id = request.POST.get("vehiculo")

        if conductor_id and vehiculo_id:
            Entrega.objects.create(
                pedido=orden,
                conductor_id=conductor_id,
                vehiculo_id=vehiculo_id
            )

            orden.estado = "en_ruta"
            orden.conductor_id = conductor_id
            orden.fecha_toma_entrega = timezone.now()
            orden.save()

            registrar_actividad(request, 'editar', 'pedidos', orden.id, f"Pedido asignado a conductor ID: {conductor_id}")

            return redirect("ordenes:lista_pedidos_admin")
        else:
            return render(request, "ordenes/asignar_entrega.html", {
                "orden": orden,
                "conductores": conductores,
                "vehiculos": vehiculos,
                "error": "Por favor selecciona un conductor y un vehículo."
            })

    return render(request, "ordenes/asignar_entrega.html", {
        "orden": orden,
        "conductores": conductores,
        "vehiculos": vehiculos
    })

@login_required
def editar_orden(request, id):
    orden = get_object_or_404(Orden, id=id)
    if request.method == "POST":
        nuevo_estado = request.POST.get("estado")
        
        if nuevo_estado == "en_ruta" and orden.estado != "en_ruta":
            orden.fecha_toma_entrega = timezone.now()
        elif nuevo_estado == "entregado" and orden.estado != "entregado":
            orden.fecha_entrega_real = timezone.now()
            
        orden.estado = nuevo_estado
        orden.save()
        registrar_actividad(request, 'editar', 'pedidos', orden.id, f"Estado de pedido cambiado a: {nuevo_estado}")
        return redirect("ordenes:lista_pedidos_admin")
    return render(request, "ordenes/detalle.html", {"orden": orden})

def format_currency(value):
    try:
        value = Decimal(value)
    except Exception:
        value = Decimal(str(value))
    formatted = f"${value:,.2f}".replace(",", ".")
    return formatted


def get_invoice_items(orden):
    items = []
    if orden.detalles.exists():
        for detalle in orden.detalles.all():
            subtotal = detalle.precio_unitario * detalle.cantidad
            items.append({
                'descripcion': detalle.material.nombre,
                'cantidad': detalle.cantidad,
                'precio_unitario': detalle.precio_unitario,
                'subtotal': subtotal
            })
    elif orden.material:
        subtotal = orden.material.precio * orden.cantidad
        items.append({
            'descripcion': orden.material.nombre,
            'cantidad': orden.cantidad,
            'precio_unitario': orden.material.precio,
            'subtotal': subtotal
        })
    else:
        items.append({
            'descripcion': 'Servicio de Transporte',
            'cantidad': orden.cantidad,
            'precio_unitario': orden.precio,
            'subtotal': orden.precio
        })
    return items


def get_invoice_totals(items, iva_rate=Decimal('0.19')):
    subtotal = sum(Decimal(item['subtotal']) for item in items)
    iva = (subtotal * iva_rate).quantize(Decimal('0.01'))
    total = (subtotal + iva).quantize(Decimal('0.01'))
    return subtotal, iva, total


def cargar_fuente(size, bold=False):
    nombre = 'arialbd.ttf' if bold else 'arial.ttf'
    try:
        return ImageFont.truetype(nombre, size)
    except Exception:
        return ImageFont.load_default()


@login_required
def descargar_factura(request, id, formato='pdf'):
    orden = get_object_or_404(Orden, id=id)

    if request.user.usuario.rol != 'admin' and orden.cliente != request.user.usuario:
        return HttpResponse('No autorizado', status=403)

    formato = formato.lower()
    items = get_invoice_items(orden)
    subtotal, iva, total = get_invoice_totals(items)
    filename = f'factura_{orden.id}.{"xlsx" if formato == "excel" else "png" if formato in ["imagen", "image", "png"] else "pdf"}'

    if formato == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Factura'

        # Colores y estilos
        title_fill = PatternFill('solid', fgColor='0f172a')
        header_fill = PatternFill('solid', fgColor='1e293b')
        info_fill = PatternFill('solid', fgColor='11203a')
        row_fill_dark = PatternFill('solid', fgColor='0b1528')
        row_fill_light = PatternFill('solid', fgColor='142038')
        total_fill = PatternFill('solid', fgColor='1e293b')
        
        title_font = Font(size=16, bold=True, color='f59e0b')
        header_font = Font(size=11, bold=True, color='FFFFFF')
        label_font = Font(size=10, bold=True, color='94a3b8')
        value_font = Font(size=10, color='FFFFFF')
        total_font = Font(size=11, bold=True, color='f59e0b')
        
        thin_border = Border(
            left=Side(style='thin', color='334155'),
            right=Side(style='thin', color='334155'),
            top=Side(style='thin', color='334155'),
            bottom=Side(style='thin', color='334155')
        )

        # Título principal (Fila 1)
        ws.merge_cells('A1:D1')
        ws['A1'] = 'Factura de Venta - Constru-Trans'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].border = thin_border
        ws.row_dimensions[1].height = 28

        # Espacio en blanco (Fila 2)
        ws.row_dimensions[2].height = 8

        # Bloc de información del cliente (Filas 3-6)
        info_data = [
            ('Factura:', f'FCT-{orden.id}'),
            ('Fecha:', orden.fecha.strftime('%Y-%m-%d %H:%M')),
            ('Cliente:', f'{orden.cliente.nombres} {orden.cliente.apellidos}'),
            ('Email:', orden.cliente.user.email)
        ]

        for row_idx, (label, value) in enumerate(info_data, start=3):
            ws[f'A{row_idx}'] = label
            ws[f'A{row_idx}'].font = label_font
            ws[f'A{row_idx}'].fill = info_fill
            ws[f'A{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'A{row_idx}'].border = thin_border

            ws.merge_cells(f'B{row_idx}:D{row_idx}')
            ws[f'B{row_idx}'] = value
            ws[f'B{row_idx}'].font = value_font
            ws[f'B{row_idx}'].fill = info_fill
            ws[f'B{row_idx}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'B{row_idx}'].border = thin_border
            ws.row_dimensions[row_idx].height = 22

        # Espacio en blanco (Fila 7)
        ws.row_dimensions[7].height = 8

        # Encabezado de tabla (Fila 8)
        headers = ['Descripción', 'Cantidad', 'Precio Unitario', 'Subtotal']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        ws.row_dimensions[8].height = 24

        # Filas de datos (Filas 9+)
        for item_idx, item in enumerate(items, start=9):
            fill = row_fill_light if (item_idx - 9) % 2 == 0 else row_fill_dark
            
            # Descripción
            desc_cell = ws.cell(row=item_idx, column=1, value=item['descripcion'])
            desc_cell.font = value_font
            desc_cell.fill = fill
            desc_cell.alignment = Alignment(horizontal='left', vertical='center')
            desc_cell.border = thin_border

            # Cantidad
            qty_cell = ws.cell(row=item_idx, column=2, value=item['cantidad'])
            qty_cell.font = value_font
            qty_cell.fill = fill
            qty_cell.alignment = Alignment(horizontal='center', vertical='center')
            qty_cell.border = thin_border

            # Precio Unitario
            price_cell = ws.cell(row=item_idx, column=3, value=float(item['precio_unitario']))
            price_cell.font = value_font
            price_cell.number_format = '$#,##0.00'
            price_cell.fill = fill
            price_cell.alignment = Alignment(horizontal='center', vertical='center')
            price_cell.border = thin_border

            # Subtotal
            subtotal_cell = ws.cell(row=item_idx, column=4, value=float(item['subtotal']))
            subtotal_cell.font = value_font
            subtotal_cell.number_format = '$#,##0.00'
            subtotal_cell.fill = fill
            subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
            subtotal_cell.border = thin_border

            ws.row_dimensions[item_idx].height = 20

        # Espacio antes de totales
        footer_start = 9 + len(items)
        ws.row_dimensions[footer_start].height = 8

        # Fila de totales
        footer_start += 1
        total_rows = [
            ('Subtotal:', float(subtotal)),
            ('IVA (19%):', float(iva)),
            ('Total:', float(total))
        ]

        for offset, (label, value) in enumerate(total_rows):
            row = footer_start + offset
            
            # Etiqueta
            label_cell = ws.cell(row=row, column=3, value=label)
            label_cell.font = total_font if label == 'Total:' else Font(size=10, bold=True, color='FFFFFF')
            label_cell.fill = total_fill
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            label_cell.border = thin_border

            # Valor
            value_cell = ws.cell(row=row, column=4, value=value)
            value_cell.font = total_font if label == 'Total:' else Font(size=10, bold=True, color='FFFFFF')
            value_cell.number_format = '$#,##0.00'
            value_cell.fill = total_fill
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            value_cell.border = thin_border

            ws.row_dimensions[row].height = 22

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

        wb.save(response)
        registrar_actividad(request, 'otro', 'pedidos', orden.id, f'Factura Excel descargada por {request.user.username}')
        return response

    if formato in ['imagen', 'image', 'png']:
        response = HttpResponse(content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        width = 1200
        line_height = 40
        header_height = 220
        body_height = 120 + (len(items) * 60)
        total_height = 260
        height = header_height + body_height + total_height

        img = Image.new('RGB', (width, height), '#0f172a')
        draw = ImageDraw.Draw(img)

        title_font = cargar_fuente(40, bold=True)
        subtitle_font = cargar_fuente(24)
        normal_font = cargar_fuente(20)
        bold_font = cargar_fuente(22, bold=True)

        draw.text((50, 40), 'Factura de Venta - Constru-Trans', font=title_font, fill='#f59e0b')
        draw.text((50, 98), f'Factura: FCT-{orden.id}', font=subtitle_font, fill='white')
        draw.text((50, 138), f'Cliente: {orden.cliente.nombres} {orden.cliente.apellidos}', font=normal_font, fill='white')
        draw.text((50, 170), f'Email: {orden.cliente.user.email}', font=normal_font, fill='white')
        draw.text((50, 202), f'Fecha: {orden.fecha.strftime("%Y-%m-%d %H:%M")}', font=normal_font, fill='white')

        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'Logo1.jpeg')
        if os.path.exists(logo_path):
            try:
                logo = Image.open(logo_path).convert('RGBA')
                logo.thumbnail((220, 120), Image.Resampling.LANCZOS if hasattr(Image, 'Resampling') else Image.ANTIALIAS)
                img.paste(logo, (width - logo.width - 50, 40), logo)
            except Exception:
                pass

        header_top = 260
        draw.rectangle([40, header_top, width - 40, header_top + 3], fill='#f59e0b')
        y = header_top + 20
        columns = [50, 450, 700, 950]
        headers = ['Descripción', 'Cantidad', 'Precio Unitario', 'Subtotal']
        for i, title in enumerate(headers):
            draw.text((columns[i], y), title, font=bold_font, fill='white')
        y += 50

        for item in items:
            draw.text((columns[0], y), str(item['descripcion']), font=normal_font, fill='white')
            draw.text((columns[1], y), str(item['cantidad']), font=normal_font, fill='white')
            draw.text((columns[2], y), format_currency(item['precio_unitario']), font=normal_font, fill='white')
            draw.text((columns[3], y), format_currency(item['subtotal']), font=normal_font, fill='white')
            y += 60

        y += 10
        draw.text((columns[2], y), 'Subtotal:', font=bold_font, fill='white')
        draw.text((columns[3], y), format_currency(subtotal), font=bold_font, fill='white')
        y += 50
        draw.text((columns[2], y), 'IVA (19%):', font=bold_font, fill='white')
        draw.text((columns[3], y), format_currency(iva), font=bold_font, fill='white')
        y += 50
        draw.text((columns[2], y), 'Total:', font=bold_font, fill='#f59e0b')
        draw.text((columns[3], y), format_currency(total), font=bold_font, fill='#f59e0b')

        output = BytesIO()
        img.save(output, format='PNG')
        output.seek(0)
        response.write(output.getvalue())
        registrar_actividad(request, 'otro', 'pedidos', orden.id, f'Factura imagen descargada por {request.user.username}')
        return response

    if formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='InvoiceTitle', fontSize=20, leading=24, alignment=1, textColor=colors.HexColor('#f59e0b')))

        elements.append(Paragraph('Factura de Venta - Constru-Trans', styles['InvoiceTitle']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f'Factura: <b>FCT-{orden.id}</b>', styles['Normal']))
        elements.append(Paragraph(f'Fecha: {orden.fecha.strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
        elements.append(Paragraph(f'Cliente: {orden.cliente.nombres} {orden.cliente.apellidos}', styles['Normal']))
        elements.append(Paragraph(f'Email: {orden.cliente.user.email}', styles['Normal']))
        elements.append(Spacer(1, 24))

        data = [[ 'Descripción', 'Cantidad', 'Precio Unitario', 'Subtotal' ]]
        for item in items:
            data.append([
                item['descripcion'],
                str(item['cantidad']),
                format_currency(item['precio_unitario']),
                format_currency(item['subtotal'])
            ])

        data.append(['', '', 'Subtotal', format_currency(subtotal)])
        data.append(['', '', 'IVA (19%)', format_currency(iva)])
        data.append(['', '', 'Total', format_currency(total)])

        table = Table(data, colWidths=[240, 100, 140, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#374151')),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#0f172a')),
            ('BACKGROUND', (2, -3), (-1, -1), colors.HexColor('#111827')),
            ('TEXTCOLOR', (2, -3), (-1, -1), colors.whitesmoke),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 16))
        elements.append(Paragraph('¡Gracias por preferir Constru-Trans!', styles['Italic']))

        doc.build(elements)
        registrar_actividad(request, 'otro', 'pedidos', orden.id, f'Factura PDF descargada por {request.user.username}')
        return response

    raise Http404('Formato no soportado')

@login_required
def eliminar_orden(request, id):
    orden = get_object_or_404(Orden, id=id)
    registrar_actividad(request, 'eliminar', 'pedidos', id, f"Pedido eliminado de cliente: {orden.cliente}")
    orden.delete()
    return redirect("ordenes:lista_pedidos_admin")
