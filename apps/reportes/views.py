from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from apps.ordenes.models import Orden
from apps.usuarios.models import Usuario, Material, Vehiculo

from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.utils.timezone import now
from apps.historial.utils import registrar_actividad
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import xml.etree.ElementTree as ET
from io import BytesIO

from apps.usuarios.views import admin_required

@admin_required
def reportes_admin(request):
    # Estadísticas de Órdenes
    ordenes = Orden.objects.all()
    total = ordenes.count()
    pendientes = ordenes.filter(estado="pendiente").count()
    en_ruta = ordenes.filter(estado="en_ruta").count()
    entregadas = ordenes.filter(estado="entregado").count()
    
    # Calcular porcentajes para la UI
    pct_pendientes = (pendientes * 100 / total) if total > 0 else 0
    pct_en_ruta = (en_ruta * 100 / total) if total > 0 else 0
    pct_entregadas = (entregadas * 100 / total) if total > 0 else 0

    # Materiales con stock crítico (< 10)
    stock_critico = Material.objects.filter(activo=True, stock_info__cantidad__lt=10).select_related('stock_info')

    context = {
        # Resumen de Órdenes
        "total_ordenes": total,
        "pendientes": pendientes,
        "en_ruta": en_ruta,
        "entregadas": entregadas,
        "pct_pendientes": pct_pendientes,
        "pct_en_ruta": pct_en_ruta,
        "pct_entregadas": pct_entregadas,
        
        # Resumen de Usuarios
        "total_clientes": Usuario.objects.filter(rol="cliente").count(),
        "total_conductores": Usuario.objects.filter(rol="conductor").count(),
        
        # Resumen de Inventario y Vehículos
        "total_materiales": Material.objects.filter(activo=True).count(),
        "total_vehiculos": Vehiculo.objects.count(),
        
        # Financiero
        "total_ingresos": ordenes.aggregate(total=Sum("precio"))["total"] or 0,

        # Stock Crítico
        "stock_critico": stock_critico,
    }
    return render(request, "reportes/lista.html", context)

@admin_required
def exportar_reporte_pdf(request, tipo):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = f"Reporte de {tipo.replace('_', ' ').capitalize()}"
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Fecha de generación: {now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 24))

    def format_money(val):
        try:
            v = float(val) / 100
            formatted = "{:,.2f}".format(v)
            return f"${formatted.replace(',', 'X').replace('.', ',').replace('X', '.')}"
        except:
            return "$0,00"

    data = []
    if tipo == 'clientes':
        data.append(['ID', 'Nombre', 'Correo', 'Teléfono', 'Estado'])
        # Optimizado con select_related('user')
        for u in Usuario.objects.filter(rol='cliente').select_related('user'):
            data.append([u.id, f"{u.nombres} {u.apellidos}", u.user.email, u.telefono, u.estado])
    
    elif tipo == 'materiales':
        data.append(['ID', 'Nombre', 'Tipo', 'Precio', 'Stock'])
        for m in Material.objects.filter(activo=True).select_related('stock_info'):
            p = m.precio or 0
            precio_formateado = format_money(p)
            data.append([m.id, m.nombre, m.tipo, precio_formateado, m.stock])

    elif tipo == 'ventas':
        data.append(['ID', 'Cliente', 'Fecha', 'Total', 'Estado'])
        # Optimizado con select_related('cliente__usuario')
        for o in Orden.objects.all().select_related('cliente__usuario'):
            p = o.precio or 0
            precio_formateado = format_money(p)
            data.append([o.id, f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}", o.fecha.strftime('%Y-%m-%d'), precio_formateado, o.estado])

    elif tipo == 'pedidos':
        data.append(['ID', 'Cliente', 'Materiales', 'Total', 'Estado'])
        # Optimizado con select_related('cliente__usuario') y prefetch_related('detalles__material')
        for o in Orden.objects.all().select_related('cliente__usuario').prefetch_related('detalles__material'):
            materiales = ", ".join([f"{d.cantidad}x {d.material.nombre}" for d in o.detalles.all()])
            p = o.precio or 0
            precio_formateado = format_money(p)
            data.append([o.id, f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}", materiales[:50] + "..." if len(materiales) > 50 else materiales, precio_formateado, o.estado])

    # Estilo de la tabla
    if data:
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("No hay datos disponibles para este reporte.", styles['Normal']))

    doc.build(elements)
    
    registrar_actividad(request, 'otro', 'reportes', None, f"Reporte de {tipo} exportado a PDF")
    
    return response

@admin_required
def exportar_reporte_excel(request, tipo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte {tipo.capitalize()}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_aligned = Alignment(horizontal="center")

    def format_money_raw(val):
        try:
            return float(val) / 100
        except:
            return 0.0

    if tipo == 'clientes':
        headers = ['ID', 'Nombre', 'Correo', 'Teléfono', 'Estado']
        ws.append(headers)
        for u in Usuario.objects.filter(rol='cliente').select_related('user'):
            ws.append([u.id, f"{u.nombres} {u.apellidos}", u.user.email, u.telefono, u.estado])
    
    elif tipo == 'materiales':
        headers = ['ID', 'Nombre', 'Tipo', 'Precio', 'Stock']
        ws.append(headers)
        for m in Material.objects.all().select_related('stock_info'):
            ws.append([m.id, m.nombre, m.tipo, format_money_raw(m.precio), m.stock])

    elif tipo == 'ventas':
        headers = ['ID', 'Cliente', 'Fecha', 'Total', 'Estado']
        ws.append(headers)
        for o in Orden.objects.all().select_related('cliente__usuario'):
            ws.append([o.id, f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}", o.fecha.strftime('%Y-%m-%d'), format_money_raw(o.precio), o.estado])

    elif tipo == 'pedidos':
        headers = ['ID', 'Cliente', 'Materiales', 'Total', 'Estado']
        ws.append(headers)
        for o in Orden.objects.all().select_related('cliente__usuario').prefetch_related('detalles__material'):
            materiales = ", ".join([f"{d.cantidad}x {d.material.nombre}" for d in o.detalles.all()])
            ws.append([o.id, f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}", materiales, format_money_raw(o.precio), o.estado])

    # Aplicar estilos a cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    registrar_actividad(request, 'otro', 'reportes', None, f"Reporte de {tipo} exportado a Excel")
    return response

@admin_required
def exportar_reporte_xml(request, tipo):
    root = ET.Element("reporte")
    root.set("tipo", tipo)
    root.set("fecha_generacion", now().strftime('%Y-%m-%d %H:%M:%S'))

    if tipo == 'clientes':
        for u in Usuario.objects.filter(rol='cliente').select_related('user'):
            item = ET.SubElement(root, "cliente")
            ET.SubElement(item, "id").text = str(u.id)
            ET.SubElement(item, "nombre").text = f"{u.nombres} {u.apellidos}"
            ET.SubElement(item, "email").text = u.user.email
            ET.SubElement(item, "telefono").text = u.telefono or ""
            ET.SubElement(item, "estado").text = u.estado

    elif tipo == 'materiales':
        for m in Material.objects.all().select_related('stock_info'):
            item = ET.SubElement(root, "material")
            ET.SubElement(item, "id").text = str(m.id)
            ET.SubElement(item, "nombre").text = m.nombre
            ET.SubElement(item, "tipo").text = m.tipo
            ET.SubElement(item, "precio").text = str(m.precio or 0)
            ET.SubElement(item, "stock").text = str(m.stock)

    elif tipo == 'ventas':
        for o in Orden.objects.all().select_related('cliente__usuario'):
            item = ET.SubElement(root, "venta")
            ET.SubElement(item, "id").text = str(o.id)
            ET.SubElement(item, "cliente").text = f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
            ET.SubElement(item, "fecha").text = o.fecha.strftime('%Y-%m-%d')
            ET.SubElement(item, "total").text = str(o.precio or 0)
            ET.SubElement(item, "estado").text = o.estado

    elif tipo == 'pedidos':
        for o in Orden.objects.all().select_related('cliente__usuario').prefetch_related('detalles__material'):
            item = ET.SubElement(root, "pedido")
            ET.SubElement(item, "id").text = str(o.id)
            ET.SubElement(item, "cliente").text = f"{o.cliente.usuario.nombres} {o.cliente.usuario.apellidos}"
            ET.SubElement(item, "total").text = str(o.precio or 0)
            ET.SubElement(item, "estado").text = o.estado
            dets = ET.SubElement(item, "detalles")
            for d in o.detalles.all():
                det = ET.SubElement(dets, "detalle")
                ET.SubElement(det, "material").text = d.material.nombre
                ET.SubElement(det, "cantidad").text = str(d.cantidad)

    # Convertir a string indentado
    from xml.dom import minidom
    xmlstr = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")

    response = HttpResponse(xmlstr, content_type='application/xml')
    response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{now().strftime("%Y%m%d")}.xml"'
    
    registrar_actividad(request, 'otro', 'reportes', None, f"Reporte de {tipo} exportado a XML")
    return response
